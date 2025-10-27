import openpyxl


def read_xlsx_book(excel_path):
    """
    this method will return the data not including the head row
    :param excel_path: the excel file path
    :return: all sheet data in workbook
    """
    book = openpyxl.load_workbook(excel_path)
    sheets = book.sheetnames
    content: list[list[list[str, int, int]]] = []

    for sheet in sheets:
        sheet_data: list = []
        # read sheet data
        sheet_obj = book[sheet]
        for row in sheet_obj.rows:
            # read cell value t1.cell(2, 1).value
            entry = [val.value for val in row]
            sheet_data.append(entry)
        sheet_data.pop(0)
        content.append(sheet_data)
    return content
